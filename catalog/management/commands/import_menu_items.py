from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand, CommandError

from catalog.menu_items_import import (
    MENU_ITEMS_CSV_DEFAULT,
    MENU_ITEMS_SHEET,
    import_menu_items,
    import_menu_items_csv,
)


class Command(BaseCommand):
    help = "Import POS menu items from csvdata/menu_items.csv (or an Excel workbook)."

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            default=MENU_ITEMS_CSV_DEFAULT,
            help=f"Path to the menu items CSV or workbook (default: {MENU_ITEMS_CSV_DEFAULT}).",
        )
        parser.add_argument(
            "--sheet",
            default=MENU_ITEMS_SHEET,
            help=f'Worksheet name for Excel imports (default: "{MENU_ITEMS_SHEET}").',
        )
        parser.add_argument(
            "--replace",
            action="store_true",
            help="Deactivate POS products not present in the import file.",
        )

    def handle(self, *args, **options):
        file_path = Path(options["file"])
        if not file_path.is_file():
            raise CommandError(f"File not found: {file_path}")

        if file_path.suffix.lower() == ".csv":
            with file_path.open("rb") as handle:
                stats = import_menu_items_csv(handle, replace=options["replace"])
        else:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet_name = options["sheet"]
            if sheet_name not in workbook.sheetnames:
                available = ", ".join(workbook.sheetnames)
                raise CommandError(f"Sheet {sheet_name!r} not found. Available: {available}")

            rows = list(workbook[sheet_name].iter_rows(values_only=True))
            stats = import_menu_items(rows, replace=options["replace"])

        self.stdout.write(self.style.SUCCESS("Menu items imported successfully."))
        for key, value in stats.items():
            self.stdout.write(f"  {key.replace('_', ' ').title()}: {value}")
