from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db.models.functions import Lower, Trim

from catalog.constants import ALL_INGREDIENT_CATEGORIES
from catalog.models import Product


def _normalize_name(value):
    if value is None:
        return ""
    return str(value).strip().casefold()


def _read_sheet_names(workbook, sheet_name):
    if sheet_name not in workbook.sheetnames:
        return []
    names = []
    for row in workbook[sheet_name].iter_rows(values_only=True):
        if not row:
            continue
        name = str(row[0]).strip() if row[0] is not None else ""
        if name:
            names.append(name)
    return names


def _match_products(names, *, ingredients_only=False):
    queryset = Product.objects.select_related("category")
    if ingredients_only:
        queryset = queryset.filter(category__name__in=ALL_INGREDIENT_CATEGORIES)
    else:
        queryset = queryset.exclude(category__name__in=ALL_INGREDIENT_CATEGORIES)

    by_name = {}
    for product in queryset.annotate(
        normalized_name=Lower(Trim("name")),
    ):
        by_name.setdefault(product.normalized_name, []).append(product)

    matched = []
    missing = []
    for raw_name in names:
        key = _normalize_name(raw_name)
        products = by_name.get(key)
        if not products:
            missing.append(raw_name)
            continue
        matched.extend(products)
    return matched, missing


class Command(BaseCommand):
    help = (
        "Mark products for daily stock take from csvdata/daily stocks.xlsx "
        "(shop stock + bakery stock sheets)."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            default="csvdata/daily stocks.xlsx",
            help="Path to the daily stocks workbook.",
        )
        parser.add_argument(
            "--clear",
            action="store_true",
            help="Clear daily_stock_take on all products before applying flags.",
        )

    def handle(self, *args, **options):
        file_path = Path(options["file"])
        if not file_path.is_file():
            raise CommandError(f"Workbook not found: {file_path}")

        try:
            import openpyxl
        except ImportError as exc:
            raise CommandError("openpyxl is required — pip install openpyxl") from exc

        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        shop_names = _read_sheet_names(workbook, "shop stock")
        bakery_names = _read_sheet_names(workbook, "bakery stock")
        workbook.close()

        if not shop_names and not bakery_names:
            raise CommandError("No product names found in shop stock or bakery stock sheets")

        if options["clear"]:
            cleared = Product.objects.filter(daily_stock_take=True).update(daily_stock_take=False)
            self.stdout.write(f"Cleared daily_stock_take on {cleared} product(s).")

        shop_products, shop_missing = _match_products(shop_names, ingredients_only=False)
        bakery_products, bakery_missing = _match_products(bakery_names, ingredients_only=True)
        product_ids = {product.id for product in shop_products + bakery_products}

        updated = Product.objects.filter(id__in=product_ids).update(daily_stock_take=True)

        self.stdout.write(
            self.style.SUCCESS(
                f"Marked {updated} product(s) for daily stock take "
                f"({len(shop_products)} shop, {len(bakery_products)} bakery ingredients)."
            )
        )

        if shop_missing:
            preview = ", ".join(shop_missing[:8])
            suffix = "…" if len(shop_missing) > 8 else ""
            self.stdout.write(
                self.style.WARNING(
                    f"Shop stock — {len(shop_missing)} name(s) not matched: {preview}{suffix}"
                )
            )
        if bakery_missing:
            preview = ", ".join(bakery_missing[:8])
            suffix = "…" if len(bakery_missing) > 8 else ""
            self.stdout.write(
                self.style.WARNING(
                    f"Bakery stock — {len(bakery_missing)} name(s) not matched: {preview}{suffix}"
                )
            )
