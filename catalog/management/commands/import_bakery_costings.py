from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand, CommandError

from catalog.bakery_import import import_bakery_costings
from catalog.kitchen_import import import_kitchen_costings

SHEET_IMPORTERS = {
    "BAKERY ": import_bakery_costings,
    "KITCHEN ": import_kitchen_costings,
}


class Command(BaseCommand):
    help = "Import products, ingredients, and recipes from CDP COSTINGS.xlsx."

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            default="CDP COSTINGS.xlsx",
            help="Path to the CDP COSTINGS workbook (default: CDP COSTINGS.xlsx).",
        )
        parser.add_argument(
            "--sheet",
            default="BAKERY ",
            help='Worksheet name (default: "BAKERY "). Also supports "KITCHEN ".',
        )

    def handle(self, *args, **options):
        file_path = Path(options["file"])
        if not file_path.is_file():
            raise CommandError(f"Workbook not found: {file_path}")

        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet_name = options["sheet"]
        if sheet_name not in workbook.sheetnames:
            available = ", ".join(workbook.sheetnames)
            raise CommandError(f"Sheet {sheet_name!r} not found. Available: {available}")

        importer = SHEET_IMPORTERS.get(sheet_name)
        if importer is None:
            supported = ", ".join(sorted(SHEET_IMPORTERS))
            raise CommandError(
                f"No importer configured for sheet {sheet_name!r}. Supported sheets: {supported}"
            )

        rows = list(workbook[sheet_name].iter_rows(values_only=True))
        stats = importer(rows)

        label = sheet_name.strip().title()
        self.stdout.write(self.style.SUCCESS(f"{label} costings imported successfully."))
        for key, value in stats.items():
            self.stdout.write(f"  {key.replace('_', ' ').title()}: {value}")
