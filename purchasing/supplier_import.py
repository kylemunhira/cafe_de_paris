import csv
import io
import re
from pathlib import Path

import openpyxl
from django.db import transaction

from catalog.costings_parse import normalize_name
from purchasing.models import Supplier

VAT_SHEET_NAMES = ("vat ", "vat")
NON_VAT_SHEET_NAMES = ("non vat", "non vat ")

CSV_HEADERS = [
    "id",
    "name",
    "vat_number",
    "address",
    "notes",
    "contact_person",
    "email",
    "phone",
    "is_active",
]


def normalize_vat_number(value):
    if value is None:
        return ""
    text = str(value).strip()
    if not text or text.lower() in ("none", "n/a", "-"):
        return ""
    if re.fullmatch(r"\d+\.0+", text):
        text = str(int(float(text)))
    return text


def supplier_lookup_key(name):
    return normalize_name(name).casefold()


def _parse_bool(value, default=True):
    if value is None or str(value).strip() == "":
        return default
    normalized = str(value).strip().lower()
    if normalized in ("true", "1", "yes", "y"):
        return True
    if normalized in ("false", "0", "no", "n"):
        return False
    raise ValueError(f"invalid is_active value: {value!r}")


def _supplier_payload_from_row(row):
    name = normalize_name(row.get("name"))
    if not name:
        raise ValueError("name is required")

    payload = {
        "name": name,
        "vat_number": normalize_vat_number(row.get("vat_number")),
        "address": normalize_name(row.get("address")),
        "notes": normalize_name(row.get("notes")),
        "contact_person": normalize_name(row.get("contact_person")),
        "email": normalize_name(row.get("email")),
        "phone": normalize_name(row.get("phone")),
    }
    if "is_active" in row and row.get("is_active") not in (None, ""):
        payload["is_active"] = _parse_bool(row.get("is_active"))
    return payload


def _upsert_supplier(payload, *, supplier_id=None, suppliers_by_id=None, suppliers_by_key=None):
    if supplier_id:
        supplier = (suppliers_by_id or {}).get(supplier_id)
        if not supplier:
            raise ValueError(f"supplier id {supplier_id} not found")
        for field, value in payload.items():
            setattr(supplier, field, value)
        supplier.save()
        return supplier, "updated"

    lookup_key = supplier_lookup_key(payload["name"])
    supplier = (suppliers_by_key or {}).get(lookup_key)
    if supplier:
        for field, value in payload.items():
            setattr(supplier, field, value)
        supplier.save()
        return supplier, "updated"

    supplier = Supplier.objects.create(**payload)
    if suppliers_by_key is not None:
        suppliers_by_key[lookup_key] = supplier
    return supplier, "created"


def _load_supplier_indexes():
    suppliers_by_id = {}
    suppliers_by_key = {}
    for supplier in Supplier.objects.all():
        suppliers_by_id[supplier.id] = supplier
        suppliers_by_key[supplier_lookup_key(supplier.name)] = supplier
    return suppliers_by_id, suppliers_by_key


def import_suppliers_rows(rows):
    created = 0
    updated = 0
    errors = []

    with transaction.atomic():
        suppliers_by_id, suppliers_by_key = _load_supplier_indexes()
        for row_number, row in enumerate(rows, start=2):
            try:
                payload = _supplier_payload_from_row(row)
                supplier_id = None
                raw_id = row.get("id")
                if raw_id not in (None, ""):
                    supplier_id = int(str(raw_id).strip())
                _, action = _upsert_supplier(
                    payload,
                    supplier_id=supplier_id,
                    suppliers_by_id=suppliers_by_id,
                    suppliers_by_key=suppliers_by_key,
                )
                if action == "created":
                    created += 1
                else:
                    updated += 1
            except Exception as exc:
                errors.append({"row": row_number, "message": str(exc)})

    if errors:
        transaction.set_rollback(True)

    return {"created": created, "updated": updated, "errors": errors}


def import_suppliers_csv(file_obj):
    try:
        text = io.TextIOWrapper(file_obj, encoding="utf-8-sig")
        reader = csv.DictReader(text)
    except UnicodeDecodeError:
        return {
            "created": 0,
            "updated": 0,
            "errors": [{"row": 0, "message": "File must be UTF-8 encoded CSV"}],
        }

    if not reader.fieldnames:
        return {
            "created": 0,
            "updated": 0,
            "errors": [{"row": 0, "message": "CSV file is empty"}],
        }

    normalized_headers = {h.strip().lower(): h for h in reader.fieldnames if h}
    if "name" not in normalized_headers:
        return {
            "created": 0,
            "updated": 0,
            "errors": [{"row": 0, "message": "Missing required column: name"}],
        }

    rows = []
    for raw_row in reader:
        row = {
            key.strip().lower(): (raw_row.get(header) or "").strip()
            for key, header in normalized_headers.items()
        }
        if not any(value for value in row.values()):
            continue
        rows.append(row)

    return import_suppliers_rows(rows)


def _parse_workbook_row(row, *, default_vat_number=""):
    if not row or not row[0]:
        return None

    name = normalize_name(row[0])
    if not name:
        return None

    description = normalize_name(row[1]) if len(row) > 1 else ""
    vat_number = default_vat_number
    if len(row) > 2 and default_vat_number == "":
        vat_number = normalize_vat_number(row[2])
    address = normalize_name(row[3]) if len(row) > 3 else ""

    return {
        "name": name,
        "vat_number": vat_number,
        "notes": description,
        "address": address,
    }


def _sheet_matches(sheet_name, candidates):
    normalized = sheet_name.strip().casefold()
    return normalized in {name.strip().casefold() for name in candidates}


def import_suppliers_workbook(file_path):
    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    rows = []

    for sheet_name in workbook.sheetnames:
        if _sheet_matches(sheet_name, VAT_SHEET_NAMES):
            sheet_rows = list(workbook[sheet_name].iter_rows(values_only=True))
            rows.extend(_parse_workbook_row(row) for row in sheet_rows[1:])
        elif _sheet_matches(sheet_name, NON_VAT_SHEET_NAMES):
            sheet_rows = list(workbook[sheet_name].iter_rows(values_only=True))
            rows.extend(
                _parse_workbook_row(row, default_vat_number="") for row in sheet_rows[1:]
            )

    parsed_rows = [row for row in rows if row]
    if not parsed_rows:
        return {
            "created": 0,
            "updated": 0,
            "errors": [{"row": 0, "message": "No supplier rows found in workbook"}],
        }

    return import_suppliers_rows(parsed_rows)


def import_suppliers_xlsx(file_obj):
    file_obj.seek(0)
    suffix = ".xlsx"
    temp_path = None
    try:
        import tempfile

        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as temp_file:
            temp_file.write(file_obj.read())
            temp_path = Path(temp_file.name)
        return import_suppliers_workbook(temp_path)
    finally:
        if temp_path and temp_path.exists():
            temp_path.unlink()


def export_suppliers_csv():
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=CSV_HEADERS)
    writer.writeheader()
    for supplier in Supplier.objects.order_by("name"):
        writer.writerow(
            {
                "id": supplier.id,
                "name": supplier.name,
                "vat_number": supplier.vat_number,
                "address": supplier.address,
                "notes": supplier.notes,
                "contact_person": supplier.contact_person,
                "email": supplier.email,
                "phone": supplier.phone,
                "is_active": "true" if supplier.is_active else "false",
            }
        )
    return output.getvalue()
